from __future__ import annotations

import csv
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from yj_studio.report import trap_report_rows, write_trap_report_csv, write_trap_report_pdf, write_trap_report_xlsx
from yj_studio.scene.layers import MeasurementLayer, TrapLayer


def _trap(name: str, rank: int, score: float) -> TrapLayer:
    return TrapLayer(
        id=f"trap-{rank}",
        name=name,
        boundary=np.asarray(
            [
                [0.0, 0.0, 10.0],
                [2.0, 0.0, 10.0],
                [2.0, 2.0, 10.0],
                [0.0, 2.0, 10.0],
                [0.0, 0.0, 10.0],
            ],
            dtype=np.float32,
        ),
        score=score,
        attributes={
            "rank": rank,
            "area_cells": 9,
            "relief_m": 12.5,
            "relief_samples": 6.25,
            "spill_level": 10.0,
            "high_inline": 1,
            "high_xline": 1,
            "source_horizon": "Top",
        },
    )


def test_trap_report_rows_are_sorted_and_summarized() -> None:
    rows = trap_report_rows([_trap("Trap-2", 2, 0.4), _trap("Trap-1", 1, 0.9)])

    assert [row["name"] for row in rows] == ["Trap-1", "Trap-2"]
    assert rows[0]["center_inline"] == 1.0
    assert rows[0]["center_xline"] == 1.0
    assert rows[0]["center_z"] == 10.0
    assert rows[0]["relief_m"] == 12.5


def test_trap_report_rows_merge_trap_evaluation_values() -> None:
    trap = _trap("Trap-1", 1, 0.9)
    evaluation = MeasurementLayer(
        name="eval",
        values={
            "GRV": 1000.0,
            "HCPV": 210.0,
            "STOIIP": 190.90909,
            "area_km2": 0.02,
            "mean_phi": 0.2,
            "gross_thickness_mean_m": 12.0,
            "cell_count": 32.0,
        },
        metadata={"trap_id": trap.id},
    )

    rows = trap_report_rows([trap], measurements=[evaluation])

    assert rows[0]["GRV"] == 1000.0
    assert rows[0]["HCPV"] == 210.0
    assert rows[0]["STOIIP"] == 190.90909
    assert rows[0]["mean_phi"] == 0.2


def test_write_trap_report_csv() -> None:
    out_dir = Path(__file__).parent / "_scratch"
    out_dir.mkdir(exist_ok=True)
    path = write_trap_report_csv([_trap("Trap-1", 1, 0.9)], out_dir / "trap_report_test.csv")

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))

    assert len(rows) == 1
    assert rows[0]["name"] == "Trap-1"
    assert rows[0]["score"] == "0.9"
    assert rows[0]["source_horizon"] == "Top"
    assert "GRV" in rows[0]


def test_write_trap_report_xlsx() -> None:
    out_dir = Path(__file__).parent / "_scratch"
    out_dir.mkdir(exist_ok=True)
    path = write_trap_report_xlsx([_trap("Trap-1", 1, 0.9)], out_dir / "trap_report_test.xlsx")

    workbook = load_workbook(path, read_only=True)
    sheet = workbook["Trap Report"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

    assert header[0] == "rank"
    assert "STOIIP" in header
    assert values[header.index("name")] == "Trap-1"
    assert values[header.index("score")] == 0.9


def test_write_trap_report_pdf() -> None:
    out_dir = Path(__file__).parent / "_scratch"
    out_dir.mkdir(exist_ok=True)
    path = write_trap_report_pdf([_trap("Trap-1", 1, 0.9)], out_dir / "trap_report_test.pdf")

    assert path.exists()
    assert path.read_bytes().startswith(b"%PDF")
