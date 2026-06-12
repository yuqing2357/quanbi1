from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font

from yj_studio.scene.layers import MeasurementLayer, TrapLayer


TRAP_REPORT_COLUMNS: tuple[str, ...] = (
    "rank",
    "name",
    "score",
    "area_cells",
    "relief_m",
    "relief_samples",
    "spill_level",
    "center_inline",
    "center_xline",
    "center_z",
    "high_inline",
    "high_xline",
    "edge_limited",
    "source_horizon",
    "GRV",
    "HCPV",
    "STOIIP",
    "area_km2",
    "mean_phi",
    "gross_thickness_mean_m",
    "cell_count",
)

TRAP_REPORT_PDF_COLUMNS: tuple[str, ...] = (
    "rank",
    "name",
    "score",
    "area_cells",
    "relief_m",
    "GRV",
    "HCPV",
    "STOIIP",
    "area_km2",
    "mean_phi",
    "cell_count",
)


def trap_report_rows(
    traps: Iterable[TrapLayer],
    *,
    measurements: Iterable[MeasurementLayer] | None = None,
) -> list[dict[str, Any]]:
    evaluation = _evaluation_by_trap_id(measurements or [])
    rows = [_trap_row(trap, evaluation.get(trap.id)) for trap in traps]
    rows.sort(key=lambda row: (int(row.get("rank") or 1_000_000), -float(row.get("score") or 0.0), row["name"]))
    return rows


def write_trap_report_csv(
    traps: Iterable[TrapLayer],
    path: str | Path,
    *,
    measurements: Iterable[MeasurementLayer] | None = None,
) -> Path:
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rows = trap_report_rows(traps, measurements=measurements)
    with out_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(TRAP_REPORT_COLUMNS))
        writer.writeheader()
        writer.writerows(rows)
    return out_path


def write_trap_report_xlsx(
    traps: Iterable[TrapLayer],
    path: str | Path,
    *,
    measurements: Iterable[MeasurementLayer] | None = None,
) -> Path:
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rows = trap_report_rows(traps, measurements=measurements)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trap Report"
    sheet.append(list(TRAP_REPORT_COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.get(column, "") for column in TRAP_REPORT_COLUMNS])
    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=len(TRAP_REPORT_COLUMNS)).coordinate[:-1]}{len(rows) + 1}"
    for col_index, column in enumerate(TRAP_REPORT_COLUMNS, start=1):
        values = [str(column), *(str(row.get(column, "")) for row in rows)]
        width = min(28, max(10, max(len(value) for value in values) + 2))
        sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = width
    workbook.save(out_path)
    return out_path


def write_trap_report_pdf(
    traps: Iterable[TrapLayer],
    path: str | Path,
    *,
    measurements: Iterable[MeasurementLayer] | None = None,
    max_rows_per_page: int = 24,
) -> Path:
    """Write a compact PDF summary table for quick review and reporting.

    CSV/XLSX keep the full column set. The PDF intentionally uses a smaller
    field subset so the table stays readable on landscape A4 pages.
    """

    from matplotlib import pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rows = trap_report_rows(traps, measurements=measurements)
    page_size = max(1, int(max_rows_per_page))
    chunks = [rows[index:index + page_size] for index in range(0, len(rows), page_size)] or [[]]

    with PdfPages(out_path) as pdf:
        for page_index, chunk in enumerate(chunks, start=1):
            fig, ax = plt.subplots(figsize=(11.69, 8.27))
            ax.axis("off")
            ax.text(
                0.01,
                0.98,
                f"Trap Report  |  traps={len(rows)}  |  page {page_index}/{len(chunks)}",
                transform=ax.transAxes,
                fontsize=13,
                fontweight="bold",
                va="top",
            )
            table = ax.table(
                cellText=[
                    [_pdf_cell(row.get(column, "")) for column in TRAP_REPORT_PDF_COLUMNS]
                    for row in chunk
                ],
                colLabels=list(TRAP_REPORT_PDF_COLUMNS),
                cellLoc="center",
                colLoc="center",
                loc="center",
                bbox=[0.01, 0.04, 0.98, 0.86],
            )
            table.auto_set_font_size(False)
            table.set_fontsize(7)
            for (row_index, _col_index), cell in table.get_celld().items():
                if row_index == 0:
                    cell.set_text_props(weight="bold")
                    cell.set_facecolor("#e8eef7")
                else:
                    cell.set_facecolor("#ffffff" if row_index % 2 else "#f7f9fc")
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)
    return out_path


def _trap_row(trap: TrapLayer, evaluation: dict[str, Any] | None = None) -> dict[str, Any]:
    attrs = dict(trap.attributes or {})
    center = _boundary_center(trap.boundary)
    values = dict(evaluation or {})
    return {
        "rank": _value(attrs, "rank", ""),
        "name": trap.name,
        "score": _round(trap.score),
        "area_cells": _value(attrs, "area_cells", ""),
        "relief_m": _round(attrs.get("relief_m")),
        "relief_samples": _round(attrs.get("relief_samples")),
        "spill_level": _round(attrs.get("spill_level")),
        "center_inline": _round(center[0]) if center is not None else "",
        "center_xline": _round(center[1]) if center is not None else "",
        "center_z": _round(center[2]) if center is not None else "",
        "high_inline": _value(attrs, "high_inline", ""),
        "high_xline": _value(attrs, "high_xline", ""),
        "edge_limited": bool(attrs.get("edge_limited", False)),
        "source_horizon": _value(attrs, "source_horizon", ""),
        "GRV": _round(values.get("GRV", attrs.get("GRV"))),
        "HCPV": _round(values.get("HCPV", attrs.get("HCPV"))),
        "STOIIP": _round(values.get("STOIIP", attrs.get("STOIIP"))),
        "area_km2": _round(values.get("area_km2", attrs.get("area_km2"))),
        "mean_phi": _round(values.get("mean_phi", attrs.get("mean_phi"))),
        "gross_thickness_mean_m": _round(values.get("gross_thickness_mean_m", attrs.get("gross_thickness_mean_m"))),
        "cell_count": _round(values.get("cell_count", attrs.get("cell_count"))),
    }


def _evaluation_by_trap_id(measurements: Iterable[MeasurementLayer]) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for layer in measurements:
        trap_id = str(layer.metadata.get("trap_id", "") or "")
        if not trap_id:
            continue
        rows[trap_id] = dict(layer.values or {})
    return rows


def _boundary_center(boundary: np.ndarray | None) -> tuple[float, float, float] | None:
    if boundary is None:
        return None
    arr = np.asarray(boundary, dtype=np.float32)
    if arr.ndim != 2 or arr.shape[0] == 0 or arr.shape[1] < 3:
        return None
    if arr.shape[0] > 1 and np.linalg.norm(arr[0, :3] - arr[-1, :3]) <= 1e-5:
        arr = arr[:-1]
    if arr.shape[0] == 0:
        return None
    center = np.nanmean(arr[:, :3], axis=0)
    return (float(center[0]), float(center[1]), float(center[2]))


def _round(value: object, digits: int = 6) -> float | str:
    if value is None:
        return ""
    try:
        return round(float(value), digits)
    except (TypeError, ValueError):
        return ""


def _value(attrs: dict[str, Any], key: str, default: Any) -> Any:
    value = attrs.get(key, default)
    return default if value is None else value


def _pdf_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.4g}"
    return str(value)
